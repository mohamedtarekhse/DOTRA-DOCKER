"""Manager pre-approval import: Excel -> whitelist vehicles + day permits.

A manager uploads an Excel sheet (columns: plate number, date) a day in
advance. Each row grants the vehicle entry AND exit for that day:

  - vehicles not yet registered are auto-created as whitelisted with
    requires_exit_permission=False (the permit window is the approval);
  - an existing vehicle is (re)activated as whitelisted the same way;
  - a whitelist_permissions row is upserted for the calendar day
    (valid_from = 00:00, valid_until = end-of-day, Africa/Cairo time).

Gate logic in gate_service only permits vehicles with an ACTIVE permission
window that contains "now", so a permit only has effect on its own day.
"""

import io
import re
from datetime import date, datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models import Vehicle, WhitelistPermission

TZ = ZoneInfo("Africa/Cairo")
UTC = timezone.utc

PLATE_KEYS = {"plate", "plate_number", "plate no", "platenumber", "plate_no", "رقم"}
DATE_KEYS = {"date", "valid", "valid_until", "valid date", "permitted on", "اليوم", "تاريخ"}
NOTES_KEYS = {"notes", "note", "remarks", "comment", "ملاحظات"}

ARABIC_DIGITS = {"٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
                 "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9"}


def normalize_plate(raw) -> str:
    """Match lpr-listener normalization: Arabic digits -> Latin, keep A-Za-z0-9 + Arabic letters (Egyptian plates)."""
    if raw is None:
        return ""
    s = "".join(ARABIC_DIGITS.get(ch, ch) for ch in str(raw))
    return re.sub(r"[^A-Za-z0-9\u0600-\u06FF]", "", s).upper()


def _find_header(cols: tuple) -> tuple[int | None, int | None, int | None]:
    plate_idx = date_idx = notes_idx = None
    for i, col in enumerate(cols):
        if col is None:
            continue
        key = str(col).strip().lower()
        if plate_idx is None and key in PLATE_KEYS:
            plate_idx = i
        elif date_idx is None and key in DATE_KEYS:
            date_idx = i
        elif notes_idx is None and key in NOTES_KEYS:
            notes_idx = i
    return plate_idx, date_idx, notes_idx


def _parse_date(value, row_no: int) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 10000:  # Excel serial date
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_workbook(data: bytes) -> tuple[list[dict], list[str]]:
    """Read the first sheet and normalize rows to {plate, date, notes}."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Not a readable .xlsx file: {exc}") from exc
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows, None)
    except StopIteration:
        raise ValueError("Excel file is empty") from None
    if not header:
        raise ValueError("Excel file is empty (no header row)")

    plate_idx, date_idx, notes_idx = _find_header(header)
    if plate_idx is None:
        name = " ,".join(str(c) for c in header if c is not None) or "(no header)"
        raise ValueError(f"Could not find a 'plate' column. Header was: {name}")
    if date_idx is None:
        raise ValueError("Could not find a 'date' column — expected columns: plate, date")

    entries, errors = [], []
    for row_no, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue
        plate = normalize_plate(row[plate_idx])
        if not plate:
            errors.append(f"Row {row_no}: missing plate")
            continue
        day = _parse_date(row[date_idx], row_no)
        if day is None:
            errors.append(f"Row {row_no}: invalid date {row[date_idx]!r}")
            continue
        entries.append({
            "plate": plate,
            "date": day,
            "notes": str(row[notes_idx]).strip() if notes_idx is not None and row[notes_idx] else None,
        })
    return entries, errors


def day_window(day: date) -> tuple[datetime, datetime]:
    """Calendar day in Africa/Cairo -> UTC aware [start, end]."""
    start_local = datetime.combine(day, time.min, tzinfo=TZ)
    end_local = datetime.combine(day + timedelta(days=1), time.min, tzinfo=TZ) - timedelta(microseconds=1)
    return start_local.astimezone(UTC), end_local.astimezone(UTC)


async def import_permits(db: AsyncSession, entries: list[dict], authorized_by: str) -> dict:
    summary = {
        "permits_created": 0,
        "permits_updated": 0,
        "vehicles_created": 0,
        "errors": [],
    }
    for entry in entries:
        plate = entry["plate"]
        valid_from, valid_until = day_window(entry["date"])
        notes = entry.get("notes")

        vehicle = (await db.execute(select(Vehicle).where(Vehicle.plate_number == plate))).scalar_one_or_none()
        if vehicle is None:
            vehicle = Vehicle(
                plate_number=plate,
                is_whitelisted=True,
                requires_exit_permission=False,  # the permit window IS the pre-approval
            )
            db.add(vehicle)
            await db.flush()
            summary["vehicles_created"] += 1
        else:
            vehicle.is_whitelisted = True
            vehicle.requires_exit_permission = False

        existing = (await db.execute(
            select(WhitelistPermission).where(
                WhitelistPermission.vehicle_id == vehicle.id,
                WhitelistPermission.valid_from == valid_from,
            )
        )).scalar_one_or_none()

        if existing is not None:
            existing.valid_until = valid_until
            existing.is_active = True
            existing.authorized_by = authorized_by
            existing.notes = notes
            summary["permits_updated"] += 1
        else:
            db.add(WhitelistPermission(
                vehicle_id=vehicle.id,
                valid_from=valid_from,
                valid_until=valid_until,
                authorized_by=authorized_by,
                is_active=True,
                notes=notes,
            ))
            summary["permits_created"] += 1

    await db.commit()
    return summary